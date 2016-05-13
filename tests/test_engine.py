import unittest
from mock import MagicMock

from openpyxl import load_workbook
from src.engine import Engine

class EngineTests(unittest.TestCase):
  def test_readsExcel(self):
    numberOfTeams = 5
    filePath = "starwars.xlsx"

    engine = Engine(numberOfTeams, filePath)

    self.assertEqual(engine.original_file.rows, 20)

  def test_usesCorrectFilePath(self):
    numberOfTeams = 5
    filePath = "starwars.xlsx"
    load_workbook._init__ = MagicMock()

    engine = Engine(numberOfTeams, filePath)

    load_workbook.__init__.assert_called_with("starwars.xlsx")

if __name__ == '__main__':
  unittest.main()
